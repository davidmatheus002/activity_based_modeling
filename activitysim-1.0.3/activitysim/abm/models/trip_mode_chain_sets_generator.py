"""
    This component will generate mode chain sets for tours with 2, 3, 4, 5, or 6 trips, assuming the agent having all
    available mode resources
"""

from activitysim.abm.models.util.trip_mode_chain_constraints_checker import *
from activitysim.core import inject
from activitysim.core import config
from .util.trip_acc_egr_choice import read_modes_from_settings, convert_mode_id2name, get_output_filename, convert_mode_name2id
from os import path
import openpyxl
import logging
import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)


def read_excel(xlsx_file_path, sheet_name, header_row_num):
    """
    read Excel information https://www.datacamp.com/community/tutorials/python-excel-tutorial#gs.somui4U
    :param xlsx_file_path:
    :param sheet_name:
    :param header_row_num: row number of the header
    :return: generated query
    """
    print('[Info]: Current workbook =', path.basename(xlsx_file_path))

    # open a workbook
    workbook = openpyxl.load_workbook(xlsx_file_path)

    # get all sheet names and compare with sheet_name
    for name in workbook.get_sheet_names():
        if sheet_name.lower() == name.lower():
            sheet_name = name
            break

    # loads current sheet
    worksheet = workbook.get_sheet_by_name(sheet_name)

    # print the sheet title with # of rows
    print("[Info]: Current worksheet =", worksheet.title, "(", worksheet.max_row, "rows)")
    last_row = worksheet.max_row

    # loop through each row (including the header row)
    for i in range(header_row_num+1, last_row+1):
        # if __debug__:
        #     debug_value = ""
        #     for cell in worksheet[i]:
        #         debug_value += " " + str(cell.value)
        #     print("[Input]:", debug_value)
        yield worksheet[i]


def read_modes(all_modes, file_full_path, tabsheet_name, use_shared_mode=False):
    """
    read available modes from Modes worksheet
    :param all_modes: dictionary of mode_id, and its property
    :param file_full_path:
    :param tabsheet_name
    :param use_shared_mode: always FALSE, indicate the whole multi-modal use shared mode, e.g. walk-car-bike it means CarShr and
    BikeShr, it is actually a personal attribute, but we can generate all the combinations in advance
    :return:
    """

    # read the Mode settings from Excel
    all_rows = read_excel(file_full_path, tabsheet_name, header_row_num=1)

    # iterate over all rows
    for row in all_rows:
        is_mode_allowed_in_scenario = 1 # row[4].value  # the "allowed" column
        mode_id = row[0].value
        if is_mode_allowed_in_scenario:
            all_modes[mode_id] = TrafficMode(L1=row[1].value, L2=row[2].value, NAME=row[3].value, allowed=1,
                                             own_mode_required=row[5].value, maas_subscription=row[6].value,
                                             use_shared_mode=use_shared_mode)


def split_chain_sets_by_main_mode(df_chain_sets, tour_type, num_trips_per_tour, filename, file_mode):
    """
    Split whole chain mode sets of a specified ownership and tour type into (main mode based)
    :param df_chain_sets:
    :param tour_type:   "2_trips", "3_trips", etc.
    :param num_trips_per_tour: number of trips in a tour
    :param filename:
    :param file_mode:
    :return:
    """
    # identify the columns: main_mode_col and trip_mode_col
    columns = {'M{}'.format(i+1): 'trip{}_mode'.format(i+1) for i in range(num_trips_per_tour)}

    # calculate the main mode
    for main_mode_col, trip_col in columns.items():
        df_chain_sets[main_mode_col] = df_chain_sets[trip_col] % 10 ** 6 // 10**3

    # # create the main mode combinations
    # mode_list = TrafficMode.mode_id_name_map.keys()
    # mode_list.remove(MODE_INVALID)
    # main_mode_combinations = list(product(mode_list, repeat=num_trips_per_tour))

    # group by main mode combinations
    local_file_mode = file_mode
    grouped = df_chain_sets.groupby(list(columns.keys()))
    logger.info("  Splitting {0} mode chain set into {1} groups.".format(len(df_chain_sets), len(grouped)))
    # counter = 0
    for name, a_group in grouped:
        # counter += 1
        # if counter % 10 == 0:
        #     print("  progress {0}/{1}".format(counter, len(grouped)), end="\r", flush=True)
        # reset index to 0-based
        # a_group.reset_index(drop=True, inplace=True)

        # save to H5
        key_name = tour_type + "/" + "_".join([str(i) for i in name])
        a_group[list(columns.values())].to_hdf(filename, key_name, mode=local_file_mode)
        local_file_mode = 'a'
    print("")


@inject.step()
def generate_trip_mode_chain_sets(settings):
    """
    Here we use sample tours from maas_modes.xlsm tabsheet 'tours_asim' and 'persons_asim' to generate the mode chain
    combinations. The combinations taken the car,bike and ebike as private ownership.
    The combinations are saved in different HDF5 files, the structure of the file is as flows:
      *_0_ownership_%N%_modes.h5:
         2_trips
         3_trips
         ...
         6_trips
         4_trips_subtour_from_2
      *_1_ownership_%N%_modes.h5: (having car)
      *_2_ownership_%N%_modes.h5: (having car and bike)
      *_3_ownership_%N%_modes.h5: (having car and ebike)
      ...
      *_7_ownership_%N%_modes.h5: (having car, bike and ebike)
    :param settings: settings in YAML file
    :return:
    """
    trace_label = 'generate_trip_mode_chain_sets'
    configs_dir = inject.get_injectable('configs_dir')
    file_name = configs_dir + "/maas_modes.xlsm"
    mode_tabsheet = 'Modes_MaaS'
    tours_tabsheet = 'tours_asim'
    persons_tabsheet = 'persons_asim'
    # those agents take typical
    selected_agent_id = range(10001, 10008+1) #, 10002, 10003, 10004, 10005, 10006, 10007]
    # selected_agent_id = [10002]  # , 10002, 10003, 10004, 10005, 10006, 10007]

    '''settings from model specific YAML files'''
    enable_maas = settings.get('enable_maas', 0)
    logger.info("Get the ENABLE_MAAS = %d from settings.yaml", enable_maas)

    # get the mode names in lower case
    multimodes_list = read_modes_from_settings(config.read_model_settings('trip_access_egress_choice.yaml'))
    logger.info("Get %d allowed mode names (lower case) from YAML file", len(multimodes_list))

    '''Data from Excel'''
    # read all multi-modal modes
    all_modes = {}  # {mode_id, TrafficMode instance}
    read_modes(all_modes, file_name, mode_tabsheet, use_shared_mode=enable_maas)
    logger.info("Get %d possible modes from " + file_name + " file", len(all_modes))
    # read sample personal information
    persons_df = pd.read_excel(file_name, sheet_name=persons_tabsheet, header=0, index_col=0)
    # read sample trips information
    trips_df = pd.read_excel(file_name, sheet_name=tours_tabsheet, header=0)

    '''ONLY use modes specified in YAML file's MNL section'''
    all_allowed_modes = {key: val for key, val in all_modes.items() if val.mode_name in multimodes_list}
    multimodes_id_list = [convert_mode_name2id(val, TrafficMode.mode_name_map) for val in multimodes_list]
    temp = [val.mode_name for key, val in all_allowed_modes.items()]
    diff = set(multimodes_list) - set(temp)
    logger.info("Finally %d modes are taken into mode chain set generation based on YAML", len(all_allowed_modes))

    # if __debug__:
    #     test_df = pd.read_hdf(output_filename, key="6_trips")
    set_created_for_when_maas_enabled = False

    '''For each kind of agent having different ownership/subscription for specified travel resource, we will 
    generate mode chain sets of each type of tour,  so iterate through each type of ownership'''
    for agent_id, agent in persons_df.groupby('agent_id'):
        if agent_id not in selected_agent_id:
            continue

        agent_s = agent.squeeze()

        '''when MaaS enabled, the agent need MaaS_subscription, make sure the agent has that! 
        The the agent can use all modes! We ONLY need to calculate 1 TIME!'''
        if set_created_for_when_maas_enabled:
            return
        if enable_maas and not set_created_for_when_maas_enabled:
            set_created_for_when_maas_enabled = True

        # setup the output file name
        output_filename = get_output_filename(agent_s[FIELD_AUTO_OWNERSHIP], agent_s[FIELD_BIKE_OWNERSHIP],
                                              agent_s[FIELD_EBIKE_OWNERSHIP], enable_maas, len(all_allowed_modes))
        file_write_mode = 'w'

        logger.info("Generating for agent type {0} ...".format(agent_id))
        # create a new HDF5 file for each type of agent
        # https://www.shanelynn.ie/select-pandas-dataframe-rows-and-columns-using-iloc-loc-and-ix/

        # iterate through each type of tour, represented by each unique tour id
        for tour_id, tour in trips_df.groupby(["tour_id"]):
            if len(tour) == 0:
                continue
            # if tour_id != 2:
            #     continue
            tour_type = tour['tour_type'].iloc[0]

            # initial the owned mode location, i.e. where is my private car or private (electric) bike
            own_mode_location = tour['origin'].iloc[0]

            # initial output: length = len(trips of the tour), they will be overwritten to save memory
            one_tour_modes_combination = [None for k in range(len(tour))]
            num_cols = len(one_tour_modes_combination)

            # create DataFrame from list with column names
            col_names = ["trip" + str(k) + "_mode" for k in range(1, len(tour) + 1)]

            trip_chain_modes_df = None
            if enable_maas:
                if num_cols <= 5:   # out of memory error if num of trips > 5
                    print(" Processing MaaS " + tour_type + "...", end=" ")
                    trip_chain_modes_df = generate_maas_combination(num_cols, col_names, multimodes_id_list)
            else:
                print(" Processing ownership {0} {1} ...".format(agent_id, tour_type), end=" ")
                # generate
                trip_chain_mode_list = []
                generate_tour_modes_options(0, agent.squeeze(), tour, all_allowed_modes, my_own_mode_name=None,
                                            outbound_mode=None, my_own_mode_location=own_mode_location,
                                            tour_modes=one_tour_modes_combination, trip_modes_sets=trip_chain_mode_list)
                # convert the list into a 2D array
                a = np.reshape(np.array(trip_chain_mode_list), (-1, num_cols))
                trip_chain_modes_df = pd.DataFrame(a, columns=col_names)

                # # map the mode integer to name
                # if __debug__:
                #     for col in col_names:
                #         trip_chain_modes_df[col + "_name"] = trip_chain_modes_df[col].apply(
                #             lambda x: convert_mode_id2name(x, TrafficMode.mode_id_name_map))
                #     # trip_chain_modes_df.iloc[:, :-len(col_names)].to_hdf(output_filename, tour_type)

            # save to Hdf5 file
            trip_chain_modes_df.iloc[:, :].to_hdf(configs_dir + '/../data/' + output_filename, tour_type, mode=file_write_mode)
            print("Generated {} mode chain set".format(len(trip_chain_modes_df.index)))

            # split this tour type based on main mode combinations and save in another file(s)
            split_chain_sets_by_main_mode(trip_chain_modes_df, tour_type, len(tour),
                                          configs_dir + '/../data/' + "split_" + output_filename, file_write_mode)

            # different tour types will be saved in the same HDF5 file of the same agent
            file_write_mode = 'a'

            if len(all_allowed_modes) == 61 and agent_id == 10008:  # this agent having car, bike and ebike
                if not enable_maas:
                    # H-W-H,  2 trips
                    if tour_id ==1:
                        assert len(trip_chain_modes_df.index) == 256    # 186 when person has no ebike
                    # H-W-S-H (H-interzone-W-H), 3 trips
                    elif tour_id == 2:
                        assert len(trip_chain_modes_df.index) == 2438   # 1777 when person has no ebike
                    # 4 trips: H - W - Shop - Escort - H
                    elif tour_id == 3:
                        assert len(trip_chain_modes_df.index) == 24232  # 18501 when person has no ebike
                    # 5 trips,
                    elif tour_id == 4:
                        assert len(trip_chain_modes_df.index) == 239152 # todo: need rerun
                    # 6 trips
                    elif tour_id == 5:
                        assert len(trip_chain_modes_df.index) == 2509926    # todo: need rerun
                    # 4 trips with subtour H-W- work sub-tour - H
                    elif tour_id == 6:
                        assert len(trip_chain_modes_df.index) == 32461  # 23111 when person has no ebike
                    else:
                        print("unknown tour_id")
                else:  # enable_maas:
                    pass
                    # # VALIDATION: available modes are CAR, EBIKE, BIKE, WPTW, WPTB, WPTC, BPTW, CPTW, walk
                    # # for "tours" tabsheet
                    # # H-W-S-H
                    # if agent_id == 10001:
                    #     assert len(trip_chain_modes_df.index) == 28231
                    # # H-W-S-H,
                    # elif agent_id == 10002 and tour_idx == 1:
                    #     assert len(trip_chain_modes_df.index) == 219
                    # # H-NM-H
                    # elif agent_id == 10002 and tour_idx == 2:
                    #     assert len(trip_chain_modes_df.index) == 39
                    # # same as 10001, the agent having car
                    # elif agent_id == 10003:
                    #     assert len(trip_chain_modes_df.index) == 218
                    # # H-W- work sub-tour - H
                    # elif agent_id == 10004:
                    #     assert len(trip_chain_modes_df.index) == 1370
                    # # H-W-H tour
                    # elif agent_id == 10005:
                    #     assert len(trip_chain_modes_df.index) == 952
                    # # H-W- 2 work sub-tours - H
                    # elif agent_id == 10006:
                    #     assert len(trip_chain_modes_df.index) == 49394
                    # # H - W - Shop - Escort - H
                    # elif agent_id == 10007:
                    #     assert len(trip_chain_modes_df.index) == 1298
                    # # H - W - subtour - Escort - H
                    # elif agent_id == 10008:
                    #     assert len(trip_chain_modes_df.index) == 7850

        logger.info("The generate mode chain sets are saved in %s", output_filename)

        # if __debug__:
        #     test = pd.read_hdf(output_filename, key="2_trips")
        #     pass
